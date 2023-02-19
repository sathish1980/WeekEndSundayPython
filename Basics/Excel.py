import openpyxl


class excelfilehandle():

    def excelread(self):
        readdata=openpyxl.load_workbook("C:\\Users\\sathishkumar\\PycharmProjects\\weekendSunday\\input\\inputdata.xlsx")
        writedata = openpyxl.load_workbook("C:\\Users\\sathishkumar\\PycharmProjects\\weekendSunday\\input\\outpudata.xlsx","w")
        sheetname = readdata["input"] # store the sheet name
        excelopenoutput = openpyxl.Workbook() # empty workbook for write
        writesheetname = excelopenoutput.active  # store the sheet name
        #cell = sheetname.cell(row=1, column=1).value
        #print(cell)
        maximumrow = sheetname.max_row
        print(maximumrow)
        maximumcolumn = sheetname.max_column
        print(maximumcolumn)
        for eachrow in range(1,maximumrow+1):
            for eachcolumn in range(1,maximumcolumn+1):
                cell = sheetname.cell(row=eachrow, column=eachcolumn).value
                writesheetname.cell(row=eachrow, column=eachcolumn).value = cell
                print(cell)
            print("\n")
        excelopenoutput.save("C:\\Users\\sathishkumar\\PycharmProjects\\weekendSunday\\input\\outpudata.xlsx")
        excelopenoutput.close()
        readdata.close()

obj = excelfilehandle()
obj.excelread()